import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
import os
import pandas as pd


# Declare second_form, third_form, and fourth_form as global variables
second_form = None
third_form = None
fourth_form = None

data_tree = None #display_registered_clients_data
treeview_visible = False  # Global variable to track the visibility of the Treeview #display_registered_clients_data
history_tree_visible = False

def login():
    username = username_entry.get()
    password = password_entry.get()

    if username == "admin" and password == "password":
        messagebox.showinfo("Login Successful", "Welcome, admin!")
       
        open_second_form()
        
    else:
        messagebox.showerror("Login Failed", "Invalid username or password.")


def open_second_form():
    global second_form, current_date  # Access the global variable
    
    # Destroy the main login form
    root.destroy()
    
    # Create the second form
    second_form = tk.Tk()
    second_form.title("(RISIS) Loan Management Systems")

    # Calculate the screen dimensions
    screen_width = second_form.winfo_screenwidth()
    screen_height = second_form.winfo_screenheight()
    second_form.geometry(f"{screen_width}x{screen_height}")
    
    def update_time_and_date():

        # Get the current date
        current_date = datetime.now()
        # Extract the day of the month
        day_of_month = current_date.day
        
        current_time = datetime.now().strftime("%H:%M:%S")
        current_date = datetime.now().strftime("%Y-%m-%d")


        # Update the labels with the current time and date
        time_value_label.config(text=current_time)
        date_value_label.config(text=current_date)

        # Schedule the function to be called after 1000ms (1 second)
        second_form.after(1000, update_time_and_date)
    

    
    # Create the group box
    group_box = tk.LabelFrame(second_form, text="Time and Date", font=("Arial", 16, "bold"))
    group_box.place(relx=0.97, rely=0.03, anchor=tk.NE)

    # Create labels to display time and date
    time_label = tk.Label(group_box, text="Current Time:", font=("Arial", 20))
    time_label.pack()
    time_value_label = tk.Label(group_box, text="", font=("Arial", 50))
    time_value_label.pack()

    date_label = tk.Label(group_box, text="Current Date:", font=("Arial", 20))
    date_label.pack()
    date_value_label = tk.Label(group_box, text="", font=("Arial", 50))
    date_value_label.pack()
    

    # Call the update_time_and_date function to start updating the time and date labels
    update_time_and_date()
    

    
    def create_and_configure_treeview(form, data):
        global data_tree
        if data_tree is not None:
            data_tree.destroy()

        data_frame = tk.Frame(form, width=400, height=300)
        data_frame.place(relx=0.495, rely=0.66, anchor="center")

        data_tree = ttk.Treeview(data_frame, columns=list(range(len(data[0]))), show="headings", height=26)
        data_tree.pack(fill=tk.BOTH, expand=True)

        column_widths = [200, 200, 150, 120, 100, 100, 100, 100, 100, 200, 100, 100, 100]  # Example column widths

        for idx, (header, width) in enumerate(zip(data[0], column_widths)):
            data_tree.heading(idx, text=header)
            data_tree.column(idx, anchor=tk.CENTER, width=width)

        for row_idx, row in enumerate(data[1:]):
            data_tree.insert("", "end", values=row)


    def toggle_treeview():
        global treeview_visible
        if not treeview_visible:  # Treeview is not visible, show it
            try:
                wb = openpyxl.load_workbook("Registered Clients.xlsx")
                sheet = wb.active
                data = list(sheet.iter_rows(values_only=True))
                create_and_configure_treeview(second_form, data)
                treeview_visible = True
                show_clients_data_btn.config(text="Hide Registered Clients Data", bg="green")
            except FileNotFoundError:
                messagebox.showerror("File Not Found", "The Registered Clients.xlsx file does not exist.")
        else:  # Treeview is visible, hide it
            data_tree.destroy()
            treeview_visible = False
            show_clients_data_btn.config(text="Show Registered Clients Data", bg="SystemButtonFace")


    def view_pending_data():
        global data_tree
        try:
            wb = openpyxl.load_workbook("Pending Clients.xlsx")
            sheet = wb.active

            # Get the data from the Excel sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            def on_row_click(event):
                # Get the index of the clicked row
                selected_row = data_tree.identify_row(event.y)

                # Remove the previous selection
                data_tree.selection_remove(data_tree.selection())

                # Highlight the clicked row
                if selected_row:
                    data_tree.selection_add(selected_row)

                # Bind the right-click event to the Treeview row
                data_tree.bind("<Button-3>", show_context_menu)

            def show_context_menu(event):
                # Create a context menu
                context_menu = tk.Menu(second_form, tearoff=0)

                # Add options (A and B) to the context menu
                context_menu.add_command(label="Confirm Loan", command=option_a_action)
                context_menu.add_command(label="Reject Loan", command=option_b_action)

                # Display the context menu at the event's x, y position
                context_menu.post(event.x_root, event.y_root)

            def option_a_action():
                selected_row = data_tree.selection()
                if selected_row:
                    index = int(selected_row[0].split("I")[1])
                    selected_data = data[index]

                    # Change the status to "Active"
                    selected_data = list(selected_data)  # Convert tuple to list to modify the data
                    status_index = data[0].index("Status")  # Find the index of the "Status" column
                    selected_data[status_index] = "Active"

                    # Save the modified data to "Registered Clients.xlsx"
                    save_to_excel(selected_data, "Registered Clients.xlsx")

                    # Select specific columns from selected_data
                    selected_columns = [selected_data[0],  # 0 column (Name)
                                        selected_data[5],  # 5 column (Amount Paid)
                                        selected_data[6],  # 6 column (Balance)
                                        selected_data[8],  # 8 column (Date Payment)
                                        selected_data[10],  # 10 column (Interest)
                                        selected_data[11],  # 11 column (Monthly Payment)
                                        selected_data[5],  # 5 column (Amount Paid)
                                        selected_data[3]]  # 3 column (Loan Amount)

                    # Save the individual client data to a new Excel file
                    name = selected_data[0]
                    filename = f"{name}.xlsx"
                    header_individual = ["Name", "Loan Amount", "Interest", "Monthly Payment", "Status", "Amount Paid", "Balance", "Date Payment"]
                    save_to_excel(selected_columns, filename, header_individual)


                    # Refresh the Treeview with updated data (remove the row)
                    data_tree.delete(selected_row)

            def option_b_action():
                selected_row = data_tree.selection()
                if selected_row:
                    index = int(selected_row[0].split("I")[1])
                    selected_data = data[index]

                    # Delete the data from "Pending Clients.xlsx"
                    delete_data_from_excel(selected_data)

                    # Display a success message
                    messagebox.showinfo("Option B Selected", f"Option B selected for:\n{selected_data}")

                    # Refresh the Treeview with updated data (remove the row)
                    data_tree.delete(selected_row)

            def delete_data_from_excel(selected_data):
                try:
                    # Load the workbook
                    wb = openpyxl.load_workbook("Pending Clients.xlsx")
                    sheet = wb.active

                    # Find the index of the selected_data in the data list
                    row_number = None
                    current_row = 2  # Start at row 2 to skip the header row
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row == selected_data:
                            row_number = current_row
                            break
                        current_row += 1

                    if row_number:
                        # Remove the selected_data row
                        sheet.delete_rows(row_number)

                        # Save the updated data back to the Excel file
                        wb.save("Pending Clients.xlsx")
                        messagebox.showinfo("Data Deleted", "Data has been deleted from Pending Clients.xlsx.")
                    else:
                        # If the selected_data is not found
                        messagebox.showerror("Data Not Found", "Selected data not found in the Pending Clients.xlsx file.")

                except FileNotFoundError:
                    messagebox.showerror("File Not Found", "The Pending Clients.xlsx file does not exist.")

            # Create a new window to display the data
            data_window = tk.Toplevel(second_form)
            data_window.title("Pending Clients")
            
            # Get the screen width and height
            screen_width = data_window.winfo_screenwidth()
            screen_height = data_window.winfo_screenheight()

            # Calculate the window position to center it on the screen
            window_width = 1600
            window_height = 600
            x = (screen_width - window_width) // 2
            y = (screen_height - window_height) // 2

            # Set the window geometry to position it at the center of the screen
            data_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

            # Create a Treeview widget to display the data in a tabular format
            data_tree = ttk.Treeview(data_window, columns=list(range(len(data[0]))), show="headings")
            data_tree.pack(fill=tk.BOTH, expand=True)

            # Define column widths for the Treeview (adjust the values as needed)
            column_widths = [200, 200, 150, 120, 100, 100, 100, 100, 100, 200, 100, 100, 100]  # Example column widths for 5 columns

            # Add column headers
            for idx, (header, width) in enumerate(zip(data[0], column_widths)):
                data_tree.heading(idx, text=header)
                data_tree.column(idx, anchor=tk.CENTER, width=width)

            # Insert the data into the Treeview
            for row_idx, row in enumerate(data[1:]):
                data_tree.insert("", "end", values=row)

            # Bind the row click event to the Treeview
            data_tree.bind("<ButtonRelease-1>", on_row_click)

            # Start the main loop for the data window
            data_window.mainloop()

        except FileNotFoundError:
            messagebox.showerror("File Not Found", "The Pending Clients.xlsx file does not exist.")




    def open_third_form():
        global second_form, third_form  # Access the global variables

        # Create the third form
        third_form = tk.Tk()
        third_form.title("Registration Form")
        
        desired_width = 800
        desired_height = 700

        # Calculate the x and y coordinates to position the form at the center of the screen
        screen_width = third_form.winfo_screenwidth()
        screen_height = third_form.winfo_screenheight()
        x_position = (screen_width - desired_width) // 2
        y_position = (screen_height - desired_height) // 2

        # Set the geometry of the form to be centered and with the desired width and height
        third_form.geometry(f"{desired_width}x{desired_height}+{x_position}+{y_position}")
        
    
        # Create a title label at the top center with bold font
        title_label = tk.Label(third_form, text="Registration Form", font=("Arial", 20, "bold"))
        title_label.pack(pady=20)
    
        
        # Create a frame to contain the input fields
        input_frame = tk.Frame(third_form)
        input_frame.pack(padx=10, pady=15)

        # Name
        name_label = tk.Label(input_frame, text="Name:", font=("Arial", 15))
        name_label.grid(row=0, column=0, sticky="w", padx=10, pady=10)
        name_entry = tk.Entry(input_frame, font=("Arial", 15), width=35, justify='center')
        name_entry.grid(row=0, column=1)

        # Address
        address_label = tk.Label(input_frame, text="Address:", font=("Arial", 15))
        address_label.grid(row=1, column=0, sticky="w", padx=10, pady=10)
        address_entry = tk.Entry(input_frame, font=("Arial", 15), width=35, justify='center')
        address_entry.grid(row=1, column=1)

        
        # Contact Number
        contact_label = tk.Label(input_frame, text="Contact Number:", font=("Arial", 15))
        contact_label.grid(row=2, column=0, sticky="w", padx=10, pady=10)
        contact_entry = tk.Entry(input_frame, font=("Arial", 15), width=35, justify='center')
        contact_entry.grid(row=2, column=1)
       
        # Date Applied
        current_date = datetime.now()
        day_of_month = current_date.day
        current_date = datetime.now().strftime("%Y-%m-%d")
        #print(day_of_month)
        #print("Current Date:", current_date)
        value_to_set = current_date
        date_applied_label = tk.Label(input_frame, text="Date Applied:", font=("Arial", 15))
        date_applied_label.grid(row=3, column=0, sticky="w", padx=10, pady=10)
        date_applied_entry = tk.Entry(input_frame, font=("Arial", 15), width=35, justify='center')
        date_applied_entry.grid(row=3, column=1)
        date_applied_entry.insert(0, value_to_set)

        
        # Type of Loan
        type_loan_label = tk.Label(input_frame, text="Type of Loan:", font=("Arial", 15))
        type_loan_label.grid(row=4, column=0, sticky="w")
        
        def set_interest_rate():
            selected_loan = type_loan_combobox.get()
            if selected_loan == "House Loan":
                interest_entry.delete(0, tk.END)
                interest_entry.insert(0, "2")
            elif selected_loan == "Money Loan":
                interest_entry.delete(0, tk.END)
                interest_entry.insert(0, "3")
            elif selected_loan == "Car Loan":
                interest_entry.delete(0, tk.END)
                interest_entry.insert(0, "4")

        type_loan_combobox = ttk.Combobox(input_frame, values=["House Loan", "Money Loan", "Car Loan"], font=("Arial", 15), width=35, justify='center')
        type_loan_combobox.grid(row=4, column=1)
        type_loan_combobox.bind("<<ComboboxSelected>>", lambda event: set_interest_rate())

        # Loan Amount
        loan_amount_label = tk.Label(input_frame, text="Loan Amount:", font=("Arial", 15))
        loan_amount_label.grid(row=5, column=0, sticky="w", padx=10, pady=10)
        loan_amount_entry = tk.Entry(input_frame, font=("Arial", 15), width=35, justify='center')
        loan_amount_entry.grid(row=5, column=1)

        # Interest
        interest_label = tk.Label(input_frame, text="Interest:", font=("Arial", 15))
        interest_label.grid(row=6, column=0, sticky="w", padx=10, pady=10)
        interest_entry = tk.Entry(input_frame, font=("Arial", 15), width=35, justify='center')
        interest_entry.grid(row=6, column=1)

        # No of Months to Pay
        no_months_pay_label = tk.Label(input_frame, text="No. of Months to Pay:", font=("Arial", 15))
        no_months_pay_label.grid(row=7, column=0, sticky="w", padx=10, pady=10)
        no_months_pay_entry = tk.Entry(input_frame, font=("Arial", 15), width=35, justify='center')
        no_months_pay_entry.grid(row=7, column=1)

        # Monthly Payment
        monthly_pay_label = tk.Label(input_frame, text="Monthly Payment:", font=("Arial", 15))
        monthly_pay_label.grid(row=8, column=0, sticky="w", padx=10, pady=10)
        monthly_pay_entry = tk.Entry(input_frame, font=("Arial", 15), width=35, justify='center')
        monthly_pay_entry.grid(row=8, column=1)

        # Requirements Presented
        requirements_label = tk.Label(input_frame, text="Requirements Presented:", font=("Arial", 15))
        requirements_label.grid(row=9, column=0, sticky="w", padx=10, pady=10)
        requirements_entry = tk.Entry(input_frame, font=("Arial", 15), width=35, justify='center')
        requirements_entry.grid(row=9, column=1)

        # Status
        status_label = ttk.Label(input_frame, text="Status:", font=("Arial", 15))
        status_label.grid(row=10, column=0, sticky="w", padx=10, pady=10)
        status_combobox = ttk.Combobox(input_frame, values=["Active", "Pending", "Cancel"], font=("Arial", 15), width=35, justify='center')
        status_combobox.grid(row=10, column=1)


        def save_data():
            name = name_entry.get()
            address = address_entry.get()
            contact_number = contact_entry.get()
            date_applied = date_applied_entry.get()
            type_loan = type_loan_combobox.get()  # Get the selected value from the combobox
            loan_amount = loan_amount_entry.get()
            interest = interest_entry.get()
            no_months_pay = no_months_pay_entry.get()
            monthly_payment = monthly_pay_entry.get()
            requirements_presented = requirements_entry.get()
            status = status_combobox.get()  # Get the selected value from the combobox
            amount_paid = "0"
            total = loan_amount

            data = [name, address, contact_number, date_applied, type_loan, loan_amount,interest, no_months_pay, monthly_payment, requirements_presented, status, amount_paid, total]
            header_individual_reg = ["Name", "Address", "Contact Number", "Date Applied", "Type of Loan", "Loan Amount", "Interest", "No. of Months to Pay", "Monthly Payment", "Requirements", "Status", "Amount Paid", "Balance"]
            
            # if status is "Active" Create a new Excel file with ID number as the filename
            if status == "Active":
          
                # Save to Registered Clients.xlsx
                save_to_excel(data, "Registered Clients.xlsx", header_individual_reg)  # Save data to main data.xlsx file
                
                totalfn = total
                amountpaidfn = "0"
                filename = f"{name}.xlsx"
                data_individual = [name, loan_amount, interest, monthly_payment, status,amountpaidfn, totalfn,date_applied]
                header_individual = ["Name", "Loan Amount", "Interest", "Monthly Payment", "Status", "Amount Paid", "Balance", "Date Payment"]
                save_to_excel(data_individual, filename, header_individual)

            # Check if the status is "Pending" and save to "Pending Clients.xlsx"
            if status == "Pending":
                save_to_excel(data, "Pending Clients.xlsx", header_individual_reg)
            
            # Clear the input fields after saving
            name_entry.delete(0, tk.END)
            address_entry.delete(0, tk.END)
            contact_entry.delete(0, tk.END)
            date_applied_entry.delete(0, tk.END)
            loan_amount_entry.delete(0, tk.END)
            interest_entry.delete(0, tk.END)
            no_months_pay_entry.delete(0, tk.END)  # Fixed the entry widget name here
            monthly_pay_entry.delete(0, tk.END)
            requirements_entry.delete(0, tk.END)

            # Reset combobox values to the first option
            type_loan_combobox.set(" ")
            status_combobox.set(" ")
            date_applied_entry.insert(0, value_to_set)
        
        # Create the "Save" button
        save_button = tk.Button(third_form, text="Save", cursor="hand2", command=save_data, font=("Arial", 15), width = 20)
        save_button.pack(pady=10)
        
  

    def save_to_excel(data, filename, header=None):
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            if header:
                sheet.append(header)

        sheet = wb.active
        sheet.append(data)
        wb.save(filename)
        messagebox.showinfo("Data Saved", f"Data has been saved to {filename}.")

    
    
    def search_data():
        # Get the name to search from the input textbox
        search_name = search_entry.get()
        
        # Check if the name exists as an Excel file in the folder
        filename = f"{search_name}.xlsx"
        if not os.path.isfile(filename):
            messagebox.showerror("File Not Found", f"No data found for {search_name}.")
            return

        try:
            # Load the workbook
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active

            # Get the data from the Excel sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            # Create Form 2
            second_form = tk.Toplevel()
            second_form.title("Search Results")
            
            
            # Label to display the file name
            filename_label = tk.Label(second_form, text=f"File Name: {filename}")
            filename_label.pack()

            def show_treeview():
                # Create a Treeview widget to display the data in a tabular format
                search_tree = ttk.Treeview(second_form, columns=list(range(len(data[0]))), show="headings")
                search_tree.pack(fill=tk.BOTH, expand=True)
                current_date = datetime.now()
                
                # Extract the day of the month
                day_of_month = current_date.day
                
          
                current_date = datetime.now().strftime("%Y-%m-%d")
                def on_row_click(event):
                    # Get the index of the clicked row
                    selected_row = search_tree.selection()
                    if selected_row:
                        selected_index = int(selected_row[0].split("I")[1])
                        selected_data = data[selected_index]

                        # Update the "Name Entry 2" widget with the name from the selected row
                        name_entry2.delete(0, tk.END)
                        name_entry2.insert(0, selected_data[0])  # Assuming the name is in the first column (index 0)
                        
                        # Update the "Loan Amount Entry 2" widget with the loan amount from the selected row
                        loan_amount_entry2.delete(0, tk.END)
                        loan_amount_entry2.insert(0, selected_data[1])  # Assuming the loan amount is in the second column (index 1)

                        # Update the "Interest Entry 2" widget with the interest from the selected row
                        interest_entry2.delete(0, tk.END)
                        interest_entry2.insert(0, selected_data[2])  # Assuming the interest is in the third column (index 2)

                        # Update the "Monthly Payment Entry 2" widget with the monthly payment from the selected row
                        monthly_payment_entry2.delete(0, tk.END)
                        monthly_payment_entry2.insert(0, selected_data[3])  # Assuming the monthly payment is in the fourth column (index 3)

                        # Update the "Status Entry 2" widget with the status from the selected row
                        status_entry2.delete(0, tk.END)
                        status_entry2.insert(0, selected_data[4])  # Assuming the status is in the fifth column (index 4)

                        # Update the "Date Paid Entry 2" widget with the date paid from the selected row
                        date_paid_entry2.delete(0, tk.END)
                        date_paid_entry2.insert(0, current_date)  # Assuming the date paid is in the sixth column (index 5)

                        # Update the "Amount Paid Entry 2" widget with the amount paid from the selected row
                        amount_paid_entry2.delete(0, tk.END)
                        amount_paid_entry2.insert(0, selected_data[5])  # Assuming the amount paid is in the seventh column (index 6)

                        # Update the "Balance Entry 2" widget with the balance from the selected row
                        balance_entry2.delete(0, tk.END)
                        balance_entry2.insert(0, selected_data[6])  # Assuming the balance is in the eighth column (index 7)

                # Bind the row click event to the Treeview
                search_tree.bind("<ButtonRelease-1>", on_row_click)

                # Add column headers
                for idx, header in enumerate(data[0]):
                    search_tree.heading(idx, text=header)
                    search_tree.column(idx, anchor=tk.CENTER)

                # Insert the data into the Treeview
                for row_idx, row in enumerate(data[1:]):
                    search_tree.insert("", "end", values=row)
            
            show_treeview()
        except FileNotFoundError:
            messagebox.showerror("File Not Found", f"The {filename} file does not exist.")
  



    
    def position_widget(widget, row, column, padx=0, pady=5, sticky="w"):
        widget.grid(row=row, column=column, padx=padx, pady=pady, sticky=sticky)


    # Create a frame to contain the input fields
    input_frame2 = tk.Frame(second_form)
    input_frame2.place(relx=0.1, rely=0.05, anchor=tk.NW) # Place at left upper position

    # Name
    name_label2 = tk.Label(input_frame2, text="Name:", font=("Arial", 15) )
    name_label2.grid(row=0, column=0, sticky="w")
    name_entry2 = tk.Entry(input_frame2, font=("Arial", 15), width=35, justify='center')
    name_entry2.grid(row=0, column=1, sticky = "w")
    
    # loan_amount
    loan_amount_label2 = tk.Label(input_frame2, text="Loan Amount:", font=("Arial", 15))
    loan_amount_label2.grid(row=1, column=0, sticky="w")
    loan_amount_entry2 = tk.Entry(input_frame2, font=("Arial", 15), width=15, justify='center')
    #loan_amount_entry2.grid(row=1, column=1)
    position_widget(loan_amount_entry2, row=1, column=1)
    
    
    # Create a frame to contain the input fields
    input_frame2_int = tk.Frame(second_form)
    input_frame2_int.place(relx=0.28, rely=0.12, anchor=tk.NW) # Place at left upper position
    # Interest
    interest_label2 = tk.Label(input_frame2_int, text="Interest:", font=("Arial", 15))
    position_widget(interest_label2, row=1, column=2)
    interest_entry2 = tk.Entry(input_frame2_int, font=("Arial", 15), width=11, justify='center')
    position_widget(interest_entry2, row=1, column=3)
    
    # Monthly Payment
    monthly_payment_label2 = tk.Label(input_frame2, text="Monthly Payment:", font=("Arial", 15))
    monthly_payment_label2.grid(row=2, column=0, sticky="w")
    monthly_payment_entry2 = tk.Entry(input_frame2, font=("Arial", 15), width=15, justify='center')
    position_widget(monthly_payment_entry2, row=2, column=1)
    
    # Status
    status_label2 = tk.Label(input_frame2_int, text="Status:", font=("Arial", 15))
    position_widget(status_label2, row=2, column=2)
    status_entry2 = tk.Entry(input_frame2_int, font=("Arial", 15), width=11, justify='center')
    position_widget(status_entry2, row=2, column=3)
    
    # Date Paid
    date_paid_label2 = tk.Label(input_frame2, text="Date Paid:", font=("Arial", 15))
    date_paid_label2.grid(row=3, column=0, sticky="w")
    date_paid_entry2 = tk.Entry(input_frame2, font=("Arial", 15), width=35, justify='center')
    position_widget(date_paid_entry2, row=3, column=1)
    
    # Amount Paid  
    amount_paid_label2 = tk.Label(input_frame2, text="  Amount Paid:", font=("Arial", 15))
    amount_paid_label2.grid(row=0, column=6, sticky="w")
    amount_paid_entry2 = tk.Entry(input_frame2, font=("Arial", 30), width=17, justify='center')
    position_widget(amount_paid_entry2, row=0, column=7)
    
    # Previous Balance
    balance_label2 = tk.Label(input_frame2, text="  Previous Balance:", font=("Arial", 15))
    balance_label2.grid(row=1, column=6, sticky="w")
    balance_entry2 = tk.Entry(input_frame2, font=("Arial", 30), width=17, justify='center')
    position_widget(balance_entry2, row=1, column=7)
    
    # Total Balance
    tot_balance_label2 = tk.Label(input_frame2, text="  Total Balance:", font=("Arial", 15))
    tot_balance_label2.grid(row=2, column=6, sticky="w")
    tot_balance_entry2 = tk.Entry(input_frame2, font=("Arial", 30), width=17, justify='center')
    position_widget(tot_balance_entry2, row=2, column=7)
    
    
    # Create a frame to contain the input fields
    input_frame_search = tk.Frame(second_form)
    input_frame_search.place(relx=0.33, rely=0.36, anchor='center')  # Place at left upper position

    # Search
    search_entry = tk.Entry(input_frame_search, font=("Arial", 15), width=40, justify='center')
    search_entry.grid(row=0, column=0, sticky="w")
    search_button = tk.Button(input_frame_search, text="Search", command=search_data, width=10)
    search_button.grid(row=0, column=1)
    
    
    show_clients_data_btn = tk.Button(input_frame_search, text="Show Registered Clients Data", command=toggle_treeview)
    show_clients_data_btn.grid(row=0, column=2)

    # Create the "View Pending Clients" button
    view_pending_button = tk.Button(input_frame_search, text="View Pending Clients", cursor="hand2", command=view_pending_data)
    view_pending_button.grid(row=0, column=6)


    # Create the "Registration" button
    registration_button = tk.Button(input_frame_search, text="Registration", cursor="hand2", command=open_third_form)
    registration_button.grid(row=0, column=5)


    def calculate_total_balance():
        try:
            # Get the values from the input fields and perform the calculation
            loan_amount = float(loan_amount_entry2.get())
            interest = float(interest_entry2.get())
            monthly_payment = float(monthly_payment_entry2.get())
            amount_paid = float(amount_paid_entry2.get())
            previous_balance = float(balance_entry2.get())
            

            total_balance = previous_balance - amount_paid
            tot_balance_entry2.delete(0, tk.END)
            tot_balance_entry2.insert(0, total_balance)
        
        except ValueError:
            tot_balance_entry2.delete(0, tk.END)
            tot_balance_entry2.insert(0, "Error: Invalid input")
    
    # Create a frame to contain the input fields
    input_frame2_cal = tk.Frame(second_form)
    input_frame2_cal.place(relx=0.52, rely=0.219, anchor=tk.NW) # Place at left upper position

    # Create the "Calculate" button
    calculate_button = tk.Button(input_frame2_cal, text="Calculate Total Balance", font=("Arial", 14), command=calculate_total_balance)
    calculate_button.grid(row=0, column=0, columnspan=2, pady=10)

    def save_to_file():
        try:
            # Get the data from the input fields
            name = name_entry2.get()
            loan_amount = loan_amount_entry2.get()
            interest = interest_entry2.get()
            monthly_payment = monthly_payment_entry2.get()
            status = status_entry2.get()
            date_paid = date_paid_entry2.get()
            amount_paid = amount_paid_entry2.get()
            total_balance = tot_balance_entry2.get()
            

            # Prepare the data as a list
            data = [name, loan_amount, interest, monthly_payment, status, amount_paid, total_balance, date_paid]

            # Save the data to a file with the name as "name.xlsx"
            filename = f"{name}.xlsx"
            header = ["Name", "Loan Amount", "Interest", "Monthly Payment", "Status", "Amount Paid", "Total Balance", "Date Paid"]
            save_to_excel(data, filename, header)

            # Append the data to the "History Payment.xlsx" file
            if os.path.exists("History Payment.xlsx"):
                wb = openpyxl.load_workbook("History Payment.xlsx")
                sheet = wb.active
            else:
                wb = openpyxl.Workbook()
                sheet = wb.active
                # Write the header row only if the file is newly created
                header = ["Name", "Loan Amount", "Interest", "Monthly Payment", "Status", "Amount Paid", "Total Balance", "Date Paid"]
                sheet.append(header)

            sheet.append(data)
            wb.save("History Payment.xlsx")

            # Clear the input fields after saving
            name_entry2.delete(0, tk.END)
            loan_amount_entry2.delete(0, tk.END)
            interest_entry2.delete(0, tk.END)
            monthly_payment_entry2.delete(0, tk.END)
            status_entry2.delete(0, tk.END)
            date_paid_entry2.delete(0, tk.END)
            amount_paid_entry2.delete(0, tk.END)
            tot_balance_entry2.delete(0, tk.END)
            balance_entry2.delete(0, tk.END)
            messagebox.showinfo("Data Saved", f"Data has been saved to {filename} and History Payment.xlsx.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving the data: {str(e)}")
     # Create the "Save" button that will handle both saving to a new Excel file and appending to "History Payment.xlsx"
    save_button2 = tk.Button(input_frame2_cal, text="Save", font=("Arial", 14), command=save_to_file)
    save_button2.grid(row=1, column=0, columnspan=2, pady=10)
    
    
    
    
    def toggle_history_treeview():
        global history_tree_visible

        if not history_tree_visible:
            show_history_payment()
            history_tree_visible = True
            show_history_button.config(text="Hide Payment History", bg="green")
        else:
            history_tree.destroy()
            history_tree_visible = False
            show_history_button.config(text="Show Payment History", bg="SystemButtonFace")
    
    # Function to show the history in a new form
    def show_history_payment():
        global history_tree
        try:
            # Load the payment history workbook and select the active sheet
            wb = openpyxl.load_workbook("History Payment.xlsx")
            sheet = wb.active

            # Get the header row and data from the sheet
            header = [cell.value for cell in sheet[1]]
            data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]

            # Create the history_frame to contain the payment history Treeview
            history_payment_frame = tk.Frame(second_form, width=400, height=300)
            history_payment_frame.place(relx=0.495, rely=0.66, anchor="center")

            # Create the payment history Treeview with the specified width and height
            
            history_tree = ttk.Treeview(history_payment_frame, columns=header, show="headings", height=26)
            history_tree.pack(fill=tk.BOTH, expand=True)

            column_widths = [250, 250, 200, 200, 200, 200, 200, 200]  # Example column widths

            for idx, (header_text, width) in enumerate(zip(header, column_widths)):
                history_tree.heading(idx, text=header_text)
                history_tree.column(idx, width=width, anchor="center")

            # Insert the data into the history Treeview
            for row in data:
                history_tree.insert("", "end", values=row)

        except FileNotFoundError:
            messagebox.showerror("File Not Found", "The History Payment.xlsx file does not exist.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while loading the data: {str(e)}")
    
    # Create the "Show History" button
    show_history_button = tk.Button(input_frame_search, text="Show History", font=("Arial", 9), command=toggle_history_treeview)
    show_history_button.grid(row=0, column=3)
    
    
    def toggle_treeview_reminder():
        global history_tree_visible

        if not history_tree_visible:
            select_clients_with_matching_last_numbers()
            history_tree_visible = True
            reminder_button.config(text="Hide Reminder Today", bg="green")
        else:
            history_tree.destroy()
            history_tree_visible = False
            reminder_button.config(text="Show Reminder Today", bg="SystemButtonFace")

    
    def select_clients_with_matching_last_numbers():
        global history_tree
        try:
            # Load the Excel file
            wb = openpyxl.load_workbook("Registered Clients.xlsx")
            sheet = wb.active

            # Get the data from the "Date Applied" column and select last numbers after the second dash
            matching_clients = []
            current_day = datetime.now().day

            for row in sheet.iter_rows(values_only=True):
                date_applied = row[3]  # Assuming "Date Applied" is in the 4th column (index 3)
                date_parts = date_applied.split("-")
                if len(date_parts) >= 3:
                    last_number = int(date_parts[2])
                    if last_number == current_day:
                        # Check if all required keys are present in the row
                        if len(row) >= 12:  # Assuming the row has at least 12 columns
                            client_data = {
                                "Name": row[0],           # Assuming client names are in the 1st column (index 0)
                                "Address": row[1],        # Assuming addresses are in the 2nd column (index 1)
                                "Contact": row[2],        # Assuming contact numbers are in the 3rd column (index 2)
                                "Date Applied": row[3],   # Assuming date applied is in the 4th column (index 3)
                                "Type Loan": row[4],      # Assuming type loan is in the 5th column (index 4)
                                "Loan Amount": row[5],    # Assuming loan amount is in the 6th column (index 5)
                                "Interest": row[6],       # Assuming interest is in the 7th column (index 6)
                                "No. of Months": row[7],  # Assuming number of months is in the 8th column (index 7)
                                "Monthly Payment": row[8],# Assuming monthly payment is in the 9th column (index 8)
                                "Requirements": row[9],  # Assuming requirements are in the 10th column (index 9)
                                "Status": row[10],       # Assuming status is in the 11th column (index 10)
                                "Amount Paid": row[11]   # Assuming amount paid is in the 12th column (index 11)
                            }
                            matching_clients.append(client_data)

            # Define column headers and widths for the Treeview (adjust the values as needed)
            column_headers = [
                "Client Name", "Address", "Contact Number", "Date Applied", "Type Loan",
                "Loan Amount", "Interest", "No. of Months", "Monthly Payment",
                "Requirements", "Status", "Amount Paid"
            ]

            column_widths = [200, 200, 150, 120, 100, 100, 100, 100, 100, 200, 100, 100]  # Example column widths for 12 columns

            # Create Treeview widget with horizontal scrollbars
            tree_frame = tk.Frame(second_form, width=400, height=300)
            tree_frame.place(relx=0.495, rely=0.66, anchor="center")


            history_tree = ttk.Treeview(tree_frame, columns=column_headers, show="headings", height=26)


            for idx, (header, width) in enumerate(zip(column_headers, column_widths)):
                history_tree.heading(idx, text=header)
                history_tree.column(idx, anchor=tk.CENTER, width=width)

            # Insert data into Treeview
            for client in matching_clients:
                history_tree.insert("", "end", values=(
                    client.get("Name", ""),
                    client.get("Address", ""),
                    client.get("Contact", ""),
                    client.get("Date Applied", ""),
                    client.get("Type Loan", ""),
                    client.get("Loan Amount", ""),
                    client.get("Interest", ""),
                    client.get("No. of Months", ""),
                    client.get("Monthly Payment", ""),
                    client.get("Requirements", ""),
                    client.get("Status", ""),
                    client.get("Amount Paid", "")
                ))

            history_tree.pack(expand=True, fill="both")


        except FileNotFoundError:
            print("The Registered Clients.xlsx file does not exist.")



        
    reminder_button = tk.Button(input_frame_search, text="Reminder Today", command=toggle_treeview_reminder)
    reminder_button.grid(row=0, column=4)
    

    def open_fourth_form():
        global fourth_form
        fourth_form = tk.Tk()  # Create a new top-level window
        fourth_form.title("Fourth Form")

        input_frame4 = tk.Frame(fourth_form)
        input_frame4.pack(padx=10, pady=10)

        # Loan Amount
        loan_amount_label = tk.Label(input_frame4, text="Loan Amount:")
        loan_amount_label.grid(row=1, column=0, sticky="w")
        loan_entry4 = tk.Entry(input_frame4)
        loan_entry4.grid(row=1, column=1)

        # Interest Rate
        interest_rate_label = tk.Label(input_frame4, text="Interest Rate:")
        interest_rate_label.grid(row=2, column=0, sticky="w")
        interest_rate_entry4 = tk.Entry(input_frame4)
        interest_rate_entry4.grid(row=2, column=1)

        # No.of Months to Pay
        number_months_label = tk.Label(input_frame4, text="Number of Months to Pay:")
        number_months_label.grid(row=3, column=0, sticky="w")
        number_months_entry = tk.Entry(input_frame4)
        number_months_entry.grid(row=3, column=1)

        def calculate4():
            # Retrieve values from input text boxes
            cal_loan_amount = float(loan_entry4.get())
            cal_interest_rate = float(interest_rate_entry4.get())
            cal_number_month = float(number_months_entry.get())

            # Interest Rate Calculation
            equal_interest_rate = (cal_loan_amount * (cal_interest_rate/100)) * cal_number_month
            # Interest Rate Label
            result_label_interest.config(text="Interest Rate: {}".format(equal_interest_rate))

            # Total Loan Amount Calculation
            equal_total_loan_amount = equal_interest_rate + cal_loan_amount
            # Total Loan Amount Label
            result_label_total_loan.config(text="Total Loan Amount: {}".format(equal_total_loan_amount))

            # Monthly Payment
            monthly_payment_rate = equal_total_loan_amount / cal_number_month
            # Monthly Payment Label
            result_label_montly_pay.config(text="Monthly Payment: {}".format(monthly_payment_rate))

        # Create the calculate button
        calculate_button = tk.Button(fourth_form, text="Calculate", command=calculate4)
        calculate_button.pack()
        result_label_interest = tk.Label(fourth_form, text="Interest Rate: ")
        result_label_interest.pack()
        result_label_total_loan = tk.Label(fourth_form, text="Total Loan Amount: ")
        result_label_total_loan.pack()
        result_label_montly_pay = tk.Label(fourth_form, text="Monthly Payment: ")
        result_label_montly_pay.pack()

        fourth_form.mainloop()    
    
        # Create Loan Calculation
    calculation_button_loan = tk.Button(input_frame_search, text="Calculation of Loan", command=open_fourth_form)
    calculation_button_loan.grid(row=0, column=7)
    
    
     # Create a frame to contain the input fields
    input_frame_credit = tk.Frame(second_form)
    input_frame_credit.place(relx=0.86, rely=0.92, anchor=tk.NW) # Place at left upper position
    
        
    # Credit label at the left lower down
    credit_label2 = tk.Label(input_frame_credit, text="ⒸLeonardo John Pansipansi and Bianca Barroa", font=("Arial", 8))
    credit_label2.grid(row=0, column=0, sticky="w")
   
    
    # Start the main loop for the second form
    second_form.mainloop()


# Create the main login form
root = tk.Tk()
root.title("(RISIS) Loan Management System")

# Calculate the screen dimensions
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate the center position for the login form
login_width = 400
login_height = 300
x_pos = (screen_width - login_width) // 2
y_pos = (screen_height - login_height) // 3

# Position the login form at the center of the screen
root.geometry(f"{login_width}x{login_height}+{x_pos}+{y_pos}")

# Styling for ttk widgets
style = ttk.Style()
style.configure("TButton", padding=10, font=("Arial", 12))
style.configure("TLabel", font=("Arial", 12))

# Create and style login form widgets
username_label = ttk.Label(root, text="Username:")
username_label.pack(pady=10)
username_entry = ttk.Entry(root, font=("Arial", 12))
username_entry.pack(pady=5)

password_label = ttk.Label(root, text="Password:")
password_label.pack()
password_entry = ttk.Entry(root, show="*", font=("Arial", 12))
password_entry.pack(pady=10)

login_button = ttk.Button(root, text="Login", command=login)
login_button.pack()

# Apply a padding around the form
login_frame = ttk.Frame(root, padding=20)
login_frame.pack()

# Credit label at the left lower down
credit_label = tk.Label(root, text="ⒸLeonardo John Pansipansi and Bianca Barroa ", font=("Arial", 8))
credit_label.pack(side=tk.RIGHT, anchor=tk.SW, padx=5, pady=10)

# Add some styling to the login form
style.configure("TFrame", background="#f0f0f0")
login_frame.configure(style="TFrame")

# Start the main loop for the login form
root.mainloop()

#pyinstaller --onefile --noconsole your_script_name.py